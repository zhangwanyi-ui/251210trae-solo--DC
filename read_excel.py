import pandas as pd
import os

# 读取Excel文件
file_path = '/Users/zhangwanyi/Desktop/陆融通达/AI生成原型/Trae测试/251210trae solo测试DC/DC页面结构272页与对应模板_251021.xlsx'

try:
    # 获取所有工作表名称
    excel_file = pd.ExcelFile(file_path)
    sheet_names = excel_file.sheet_names
    
    print(f"Excel文件包含以下工作表：")
    for sheet in sheet_names:
        print(f"- {sheet}")
    
    # 遍历所有工作表
    for sheet_name in sheet_names:
        print(f"\n=== 工作表：{sheet_name} ===")
        
        # 读取工作表数据
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        
        # 打印数据的基本信息
        print(f"数据形状：{df.shape}")
        print(f"列名：{list(df.columns)}")
        
        # 检查是否包含模板相关列
        template_columns = [col for col in df.columns if '模板' in col or 'Template' in col or '类型' in col]
        if template_columns:
            print(f"\n模板相关列：{template_columns}")
            
            # 打印模板列的唯一值
            for col in template_columns:
                unique_values = df[col].dropna().unique()
                print(f"\n{col}的唯一值 ({len(unique_values)} 个)：")
                for val in unique_values:
                    print(f"  - {val}")
                    
                    # 统计每个模板对应的页面数量
                    template_count = len(df[df[col] == val])
                    print(f"    对应页面数量：{template_count}")
        else:
            # 打印前几行数据查看内容
            print(f"\n前5行数据：")
            print(df.head())
            
            # 检查是否有其他方式标识模板
            print(f"\n数据值示例：")
            for col in df.columns[:5]:  # 只检查前5列
                unique_vals = df[col].dropna().unique()[:3]  # 只显示前3个唯一值
                print(f"  {col}: {unique_vals}")
        
except Exception as e:
    print(f"读取Excel文件时出错：{e}")
    print("尝试使用openpyxl库...")
    
    # 尝试使用openpyxl
    try:
        from openpyxl import load_workbook
        
        workbook = load_workbook(file_path)
        sheets = workbook.sheetnames
        
        print(f"\n使用openpyxl读取到的工作表：")
        for sheet in sheets:
            print(f"- {sheet}")
            
            # 读取工作表
            worksheet = workbook[sheet]
            
            # 打印前10行数据
            print(f"\n{sheet}工作表前10行数据：")
            for row in worksheet.iter_rows(min_row=1, max_row=10, values_only=True):
                # 只打印非空值
                non_empty_values = [cell for cell in row if cell is not None]
                if non_empty_values:
                    print(non_empty_values)
                    
    except Exception as e2:
        print(f"使用openpyxl读取时出错：{e2}")
        print("\n请确保已安装pandas和openpyxl库。")