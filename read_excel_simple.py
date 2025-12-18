# 简单的Excel读取脚本，只使用openpyxl库

try:
    from openpyxl import load_workbook
    import os
    
    # 读取Excel文件
    file_path = '/Users/zhangwanyi/Desktop/陆融通达/AI生成原型/Trae测试/251210trae solo测试DC/DC页面结构272页与对应模板_251021.xlsx'
    
    print(f"正在读取文件: {file_path}")
    
    # 加载工作簿
    wb = load_workbook(file_path)
    
    # 获取所有工作表名称
    sheet_names = wb.sheetnames
    print(f"\n文件包含 {len(sheet_names)} 个工作表:")
    for i, sheet in enumerate(sheet_names):
        print(f"{i+1}. {sheet}")
    
    # 遍历所有工作表
    for sheet_name in sheet_names:
        print(f"\n\n=== 工作表: {sheet_name} ===")
        
        # 获取当前工作表
        ws = wb[sheet_name]
        
        # 打印工作表的尺寸
        print(f"工作表尺寸: {ws.dimensions}")
        
        # 打印前10行的前5列数据
        print("\n前10行数据 (只显示非空单元格):")
        for row in ws.iter_rows(min_row=1, max_row=10, max_col=5, values_only=True):
            # 过滤掉空行
            if any(cell is not None for cell in row):
                # 只显示非空单元格
                non_empty_cells = [cell for cell in row if cell is not None]
                print(non_empty_cells)
        
        # 尝试查找模板相关内容
        print("\n搜索模板相关内容:")
        found_templates = False
        for row in ws.iter_rows(values_only=True):
            if any(cell and "模板" in str(cell) for cell in row):
                found_templates = True
                non_empty_cells = [cell for cell in row if cell is not None]
                print(non_empty_cells)
        
        if not found_templates:
            print("未在当前工作表找到模板相关内容")
    
    print("\n\n=== 分析完成 ===")
    print("请根据上述输出查看Excel文件的结构和内容，特别是查找包含'模板'字样的行。")
    
except Exception as e:
    print(f"读取Excel文件时出错: {e}")
    print("\n尝试安装openpyxl库...")
    
    # 尝试安装openpyxl
    try:
        import subprocess
        subprocess.run(["pip3", "install", "openpyxl", "--timeout", "10"])
        print("openpyxl安装成功，重新运行脚本即可读取Excel文件。")
    except Exception as e2:
        print(f"安装openpyxl失败: {e2}")
        print("\n请手动安装openpyxl库，然后重新运行脚本。")
        print("安装命令: pip3 install openpyxl")